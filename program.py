from openpyxl import load_workbook
import aspose.words as aw
from varname import nameof

wb = load_workbook('initial.xlsx')

ws = wb.active

project_name = str(ws['B1'].value)
location = str(ws['B2'].value)
investor = str(ws['B3'].value)
foundation = str(ws['B4'].value)
found_size = str(ws['B5'].value)
stories_number = str(ws['B6'].value)
columns_number = str(ws['B7'].value)
roof_type = str(ws['B8'].value)
snow = str(ws['B9'].value)
eq_intensity = str(ws['B10'].value)
eq_coef = str(ws['B11'].value)
region = str(ws['B12'].value)
drawings_number = str(ws['B13'].value)

# variables_dict = {nameof(project_name):project_name ,nameof(location):location, nameof(investor):investor, nameof(foundation):foundation, 
# nameof(found_size):found_size, nameof(stories_number):stories_number, nameof(columns_number):columns_number, nameof(roof_type):roof_type, nameof(snow):snow,
# nameof(eq_intensity):eq_intensity, nameof(eq_coef):eq_coef}



# ToDo: use a for loop 

# for variable in variables_list:
#     variable_name = nameof(variable)
#     document.range.replace(variable_name, variable, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))

# def replace_string(var):
#     variable_name = nameof(var)
#     print(variable_name)

#     document.range.replace(variable_name, var, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))

template_file = ('Template_Zapiska.docx')
document = aw.Document(template_file)

def replace_placeholders(document):

    document.range.replace("project_name", project_name, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("location", location, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("investor", investor, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("foundation", foundation, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("found_size", found_size, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("stories_number", stories_number, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("columns_number", columns_number, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("roof_type", roof_type, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("snow", snow, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("eq_intensity", eq_intensity, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("eq_coef", eq_coef, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("region", region, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))
    document.range.replace("drawings_number", drawings_number, aw.replacing.FindReplaceOptions(aw.replacing.FindReplaceDirection.FORWARD))

    
replace_placeholders(document)

document.save(f"Zapiska_{investor}.docx")


template_file = ('Template_TK.docx')
document = aw.Document(template_file)

replace_placeholders(document)

document.save(f"Ocenka_Suotvetstvie_{investor}.docx")



